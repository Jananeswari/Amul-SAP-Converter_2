"""
ETL / Conversion Engine
Joins platform order data against the master mapping, converts ordered
units into Amul SAP box/case quantities (rounding up on any remainder
and flagging it), and builds the manager-facing monthly projection
output with one Qty + Date column pair per platform.
"""

import math
import pandas as pd
from mapping_engine import load_master_mapping, clean_sku_value


def convert_platform_orders(platform_orders: pd.DataFrame, platform: str) -> dict:
    """
    Maps one platform's normalized order rows to SAP codes and computes
    both case and piece quantities, regardless of which the platform's
    own file originally supplied.

    Two cases:
      - Platform supplies PIECES only (Zepto, Swiggy, Blinkit): pieces /
        case_pack_size = cases, rounded UP, with any remainder flagged
        for manager review.
      - Platform supplies CASES directly (Reliance): cases is used as-is
        with no rounding (nothing to round — it's already whole cases),
        and pieces = cases x case_pack_size is back-filled for display.

    Returns dict with: mapped (df), unmapped (df), stats.
    """
    platform_orders = platform_orders.copy()
    platform_orders["platform_sku"] = platform_orders["platform_sku"].apply(clean_sku_value)

    mapping = load_master_mapping()
    plat_map = mapping[mapping["platform"] == platform][
        ["platform_sku", "sap_code", "sap_description", "fg_group",
         "case_pack_size", "pack_confidence"]
    ]

    merged = platform_orders.merge(plat_map, on="platform_sku", how="left")

    is_mapped = merged["sap_code"].notna()
    mapped = merged[is_mapped].copy()
    unmapped = merged[~is_mapped].copy()

    if len(mapped):
        mapped["case_pack_size"] = mapped["case_pack_size"].fillna(1)

        supplies_cases_directly = (
            "order_qty_cases" in mapped.columns
            and mapped["order_qty_cases"].notna().any()
        )

        if supplies_cases_directly:
            # Reliance-style: cases given directly, no rounding needed.
            # Pieces back-filled for display only.
            mapped["po_qty_cases"] = pd.to_numeric(mapped["order_qty_cases"], errors="coerce").fillna(0)
            mapped["po_qty_pieces"] = mapped["po_qty_cases"] * mapped["case_pack_size"]
            mapped["had_remainder"] = False  # nothing to round; cases are authoritative
        else:
            # Zepto/Swiggy/Blinkit-style: pieces given, derive cases.
            mapped["po_qty_pieces"] = mapped["order_qty_units"]
            raw_cases = mapped["order_qty_units"] / mapped["case_pack_size"]
            mapped["po_qty_cases"] = raw_cases.apply(math.ceil)
            mapped["had_remainder"] = (mapped["order_qty_units"] % mapped["case_pack_size"]) != 0

        mapped["platform"] = platform

    stats = {
        "platform": platform,
        "total_rows": len(merged),
        "mapped_rows": len(mapped),
        "unmapped_rows": len(unmapped),
        "rounded_up_rows": int(mapped["had_remainder"].sum()) if len(mapped) else 0,
        "total_units": float(platform_orders["order_qty_units"].sum()),
        "total_boxes": float(mapped["po_qty_cases"].sum()) if len(mapped) else 0.0,
    }
    return {"mapped": mapped, "unmapped": unmapped, "stats": stats}


def find_duplicate_platform_codes(all_mapped: dict) -> dict:
    """
    Detects, per platform, any platform_sku value that appears mapped to
    MORE THAN ONE different SAP Code. This usually means the master file
    has the same platform code listed against two different products (or
    two different pack sizes of the same product) — a real mapping issue
    the manager should check before trusting either quantity.

    Returns {platform_name: set of conflicting platform_sku values}.
    Only includes platforms/codes where a genuine conflict exists.
    """
    conflicts = {}
    for platform, df in all_mapped.items():
        if df.empty:
            continue
        counts = df.groupby("platform_sku")["sap_code"].nunique()
        conflicting_skus = set(counts[counts > 1].index)
        if conflicting_skus:
            conflicts[platform] = conflicting_skus
    return conflicts


def build_manager_projection(all_mapped: dict) -> pd.DataFrame:
    """
    Builds the manager-facing wide table, one row per SAP Code:

      FG Group | SAP Code | SAP Product Description |
      [for each platform:] Order Qty (Cases) | Order Qty (Pieces) |
                            Platform Code | Platform Description |
                            Rounded Up? | Order Date
      Total PO Qty (Cases)

    Where a platform did not order a given SAP Code, that platform's
    columns are left blank (not 0) for that row.

    If the SAME SAP Code appears on multiple order lines from the SAME
    platform (e.g. several lots/SKUs all mapping to one SAP Code),
    quantities are summed and platform code/description show the first
    one seen, since a single cell can't hold multiple raw SKUs.

    all_mapped: dict of {platform_name: mapped_dataframe}
    """
    base_cols = ["sap_code", "sap_description", "fg_group"]
    platform_frames = []

    for platform, df in all_mapped.items():
        if df.empty:
            continue
        grp = (
            df.groupby(base_cols, as_index=False)
            .agg(
                po_qty_cases=("po_qty_cases", "sum"),
                po_qty_pieces=("po_qty_pieces", "sum"),
                platform_sku=("platform_sku", "first"),
                raw_product_name=("raw_product_name", "first"),
                order_date=("order_date", "max"),
                had_remainder=("had_remainder", "any"),
            )
        )
        grp = grp.rename(columns={
            "po_qty_cases":      f"{platform} Order Qty (Cases)",
            "po_qty_pieces":     f"{platform} Order Qty (Pieces)",
            "platform_sku":      f"{platform} Platform Code",
            "raw_product_name":  f"{platform} Platform Description",
            "order_date":        f"{platform} Order Date",
            "had_remainder":     f"{platform} Rounded Up",
        })
        platform_frames.append(grp)

    if not platform_frames:
        return pd.DataFrame()

    result = platform_frames[0]
    for pf in platform_frames[1:]:
        result = result.merge(pf, on=base_cols, how="outer")

    case_cols = [c for c in result.columns if c.endswith("Order Qty (Cases)")]
    result["Total PO Qty (Cases)"] = result[case_cols].fillna(0).sum(axis=1)

    result = result.rename(columns={
        "sap_code": "SAP Code",
        "sap_description": "SAP Product Description",
        "fg_group": "FG Group",
    })

    # Column order: FG Group, SAP Code, SAP Description, then each
    # platform's full column group in upload order, then the cases total.
    ordered_cols = ["FG Group", "SAP Code", "SAP Product Description"]
    for platform in all_mapped.keys():
        for suffix in ["Order Qty (Cases)", "Order Qty (Pieces)", "Platform Code",
                       "Platform Description", "Rounded Up", "Order Date"]:
            col = f"{platform} {suffix}"
            if col in result.columns:
                ordered_cols.append(col)
    ordered_cols.append("Total PO Qty (Cases)")
    ordered_cols = [c for c in ordered_cols if c in result.columns]

    final = result[ordered_cols].sort_values("SAP Code").reset_index(drop=True)

    # Hidden helper column: True if THIS row's SAP Code is one where at
    # least one platform's code on this row is a platform_sku that maps
    # to more than one SAP Code elsewhere in that platform's data (point 5).
    # Not part of the visible column order — callers use it for highlighting.
    duplicate_codes = find_duplicate_platform_codes(all_mapped)
    if duplicate_codes:
        def row_has_conflict(row):
            for platform, conflicting_skus in duplicate_codes.items():
                code_col = f"{platform} Platform Code"
                if code_col in row and pd.notna(row[code_col]) and row[code_col] in conflicting_skus:
                    return True
            return False
        final["_has_code_conflict"] = final.apply(row_has_conflict, axis=1)
    else:
        final["_has_code_conflict"] = False

    return final


def export_projection_to_excel(projection_df: pd.DataFrame, unmapped_by_platform: dict,
                                stats_by_platform: list) -> bytes:
    """Builds the final downloadable Excel with Summary, Projection (rounded
    rows highlighted yellow, duplicate-platform-code rows highlighted light
    red), and Unmapped SKUs sheets."""
    import io
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    has_conflict_col = "_has_code_conflict" in projection_df.columns
    conflict_flags = projection_df["_has_code_conflict"].tolist() if has_conflict_col else []
    visible_df = projection_df.drop(columns=["_has_code_conflict"]) if has_conflict_col else projection_df

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(stats_by_platform).to_excel(writer, sheet_name="Summary", index=False)
        visible_df.to_excel(writer, sheet_name="Monthly PO Projection", index=False)

        unmapped_frames = []
        for platform, df in unmapped_by_platform.items():
            if not df.empty:
                tmp = df[["platform_sku", "raw_product_name", "order_qty_units"]].copy()
                tmp.insert(0, "platform", platform)
                unmapped_frames.append(tmp)
        if unmapped_frames:
            pd.concat(unmapped_frames, ignore_index=True).to_excel(
                writer, sheet_name="Unmapped SKUs", index=False
            )

        ws = writer.sheets["Monthly PO Projection"]
        yellow = PatternFill("solid", start_color="FFF6CC", end_color="FFF6CC")
        bold_orange = Font(color="B45309", bold=True)
        light_red = PatternFill("solid", start_color="FADBD8", end_color="FADBD8")
        header = [c.value for c in ws[1]]
        col_letter_by_name = {h: get_column_letter(i) for i, h in enumerate(header, start=1) if h}

        # Light red takes priority: if a row has a duplicate-platform-code
        # conflict (point 5), highlight the WHOLE row and skip the
        # yellow "rounded up" cell-level highlight for that row, since the
        # conflict is the more serious thing for the manager to check first.
        for row_idx in range(2, ws.max_row + 1):
            df_row_idx = row_idx - 2
            if has_conflict_col and df_row_idx < len(conflict_flags) and conflict_flags[df_row_idx]:
                for col_idx in range(1, len(header) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = light_red

        # Highlight "Rounded Up" = True cells and their matching
        # "Order Qty (Cases)" cell for the SAME platform, found by name
        # rather than a fixed column offset (robust to column reordering).
        # Skipped for rows already flagged with the light-red conflict above.
        for h in header:
            if not h or "Rounded Up" not in h:
                continue
            platform_prefix = h.rsplit(" Rounded Up", 1)[0]
            qty_col_name = f"{platform_prefix} Order Qty (Cases)"
            if qty_col_name not in col_letter_by_name:
                continue
            col_letter_flag = col_letter_by_name[h]
            col_letter_qty = col_letter_by_name[qty_col_name]
            for row in range(2, ws.max_row + 1):
                df_row_idx = row - 2
                if has_conflict_col and df_row_idx < len(conflict_flags) and conflict_flags[df_row_idx]:
                    continue  # already light-red highlighted, don't override
                val = ws[f"{col_letter_flag}{row}"].value
                if val is True:
                    ws[f"{col_letter_qty}{row}"].fill = yellow
                    ws[f"{col_letter_qty}{row}"].font = bold_orange

        # Auto column width across all sheets
        for sheet in writer.sheets.values():
            for col in sheet.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    return output.getvalue()
